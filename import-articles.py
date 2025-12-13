#!/usr/bin/env python3
"""
Import blog articles from Excel file to database
"""
import openpyxl
import mysql.connector
import os
import re
from datetime import datetime
from html.parser import HTMLParser

class HTMLTextExtractor(HTMLParser):
    """Extract text content from HTML"""
    def __init__(self):
        super().__init__()
        self.text = []
    
    def handle_data(self, data):
        self.text.append(data)
    
    def get_text(self):
        return ''.join(self.text).strip()

def extract_text_from_html(html_content):
    """Extract plain text from HTML"""
    parser = HTMLTextExtractor()
    parser.feed(str(html_content))
    return parser.get_text()

def clean_html_content(content):
    """Clean HTML content by removing DOCTYPE and html/head/body tags"""
    if not content:
        return ""
    
    content = str(content)
    
    # Remove DOCTYPE
    content = re.sub(r'<!DOCTYPE[^>]*>', '', content, flags=re.IGNORECASE)
    
    # Extract body content if exists
    body_match = re.search(r'<body[^>]*>(.*?)</body>', content, re.DOTALL | re.IGNORECASE)
    if body_match:
        content = body_match.group(1)
    else:
        # Remove html, head, body tags
        content = re.sub(r'</?html[^>]*>', '', content, flags=re.IGNORECASE)
        content = re.sub(r'<head>.*?</head>', '', content, flags=re.DOTALL | re.IGNORECASE)
        content = re.sub(r'</?body[^>]*>', '', content, flags=re.IGNORECASE)
    
    return content.strip()

# Database connection
db_url = os.getenv('DATABASE_URL')
if not db_url:
    print("ERROR: DATABASE_URL environment variable not set")
    exit(1)

# Parse DATABASE_URL (format: mysql://user:pass@host:port/dbname)
match = re.match(r'mysql://([^:]+):([^@]+)@([^:]+):(\d+)/(.+)', db_url)
if not match:
    print("ERROR: Invalid DATABASE_URL format")
    exit(1)

user, password, host, port, database = match.groups()

print(f"Connecting to database: {host}:{port}/{database}")

conn = mysql.connector.connect(
    host=host,
    port=int(port),
    user=user,
    password=password,
    database=database
)
cursor = conn.cursor()

# Load Excel file
print("\nLoading Excel file...")
wb = openpyxl.load_workbook('/home/ubuntu/upload/car_insurance_articles_100_images.xlsx')
sheet = wb.active

print(f"Found {sheet.max_row - 1} articles to import\n")

# Import articles
imported = 0
skipped = 0
errors = 0

for row_num in range(2, sheet.max_row + 1):
    try:
        row = list(sheet[row_num])
        
        title = row[0].value
        slug = row[1].value
        content = clean_html_content(row[2].value)
        excerpt_raw = row[3].value
        category = row[4].value
        tags = row[5].value
        meta_title = row[6].value
        meta_description_raw = row[7].value
        image_filename = row[8].value
        
        # Extract text from HTML for excerpt and meta description
        excerpt = extract_text_from_html(excerpt_raw) if excerpt_raw else ""
        meta_description = extract_text_from_html(meta_description_raw) if meta_description_raw else ""
        
        # Truncate if too long
        if len(excerpt) > 500:
            excerpt = excerpt[:497] + "..."
        if len(meta_description) > 160:
            meta_description = meta_description[:157] + "..."
        if len(meta_title) > 60:
            meta_title = meta_title[:60]
        
        # Build image URL
        cover_image = f"/images/blog/{image_filename}" if image_filename else ""
        
        # Check if article already exists
        cursor.execute("SELECT id FROM blog_posts WHERE slug = %s", (slug,))
        if cursor.fetchone():
            print(f"⏭️  Skipping (exists): {title[:50]}...")
            skipped += 1
            continue
        
        # Insert article
        published_at = datetime.now()
        
        cursor.execute("""
            INSERT INTO blog_posts 
            (title, slug, content, excerpt, cover_image, category, tags, 
             meta_title, meta_description, status, published_at, created_at, updated_at)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            title, slug, content, excerpt, cover_image, category, tags,
            meta_title, meta_description, 'published', published_at, 
            published_at, published_at
        ))
        
        conn.commit()
        print(f"✅ Imported: {title[:60]}...")
        imported += 1
        
    except Exception as e:
        print(f"❌ Error importing row {row_num}: {str(e)[:100]}")
        errors += 1
        conn.rollback()

cursor.close()
conn.close()

print(f"\n{'='*60}")
print(f"Import completed!")
print(f"✅ Imported: {imported}")
print(f"⏭️  Skipped: {skipped}")
print(f"❌ Errors: {errors}")
print(f"{'='*60}")
